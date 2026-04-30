from flask import Flask, render_template, jsonify, request, session, redirect, url_for
import sqlite3, os, hashlib, re
from functools import wraps
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'insightmrm-secret-2024-xK9pQ'
DB = "mrm.db"
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def sanitize(col):
    col = str(col)
    col = col.replace(' ','_').replace('(','').replace(')','').replace('#','num')
    col = col.replace('/','_').replace('-','_').replace('%','pct').replace('.','').replace(',','')
    return re.sub(r'_+','_',col).strip('_')

def db_has_data():
    try:
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM model_inventory")
        count = c.fetchone()[0]; conn.close()
        return count > 0
    except: return False

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session: return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('role') != 'admin': return render_template('403.html'), 403
        return f(*args, **kwargs)
    return decorated

def privileged_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('role') not in ('admin', 'developer', 'dev'): return render_template('403.html'), 403
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM users WHERE username=? AND is_active=1", (username,))
        user = c.fetchone(); conn.close()
        if user and user['password_hash'] == hash_pw(password):
            session['user_id']=user['id']; session['username']=user['username']
            session['role']=user['role']; session['full_name']=user['full_name'] or username
            return redirect(url_for('model_overview'))
        return render_template('login.html', error='Invalid username or password')
    if 'user_id' in session: return redirect(url_for('model_overview'))
    return render_template('login.html')

@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))

@app.route('/')
@login_required
def index(): return redirect(url_for('model_overview'))

@app.route('/model-overview')
@login_required
def model_overview(): return render_template('model_overview.html')

@app.route('/model-lifecycle')
@login_required
def model_lifecycle(): return render_template('model_lifecycle.html')

@app.route('/model-inventory')
@login_required
def model_inventory(): return render_template('model_inventory.html')

@app.route('/model-validation')
@login_required
def model_validation(): return render_template('model_validation.html')

@app.route('/model-findings')
@login_required
def model_findings(): return render_template('model_findings.html')

@app.route('/executive-summary')
@login_required
def executive_summary(): return render_template('executive_summary.html')

@app.route('/timeline')
@login_required
def timeline(): return render_template('timeline.html')

@app.route('/validation-checklist')
@login_required
def validation_checklist(): return render_template('validation_checklist.html')

@app.route('/upload')
@admin_required
def upload_page(): return render_template('upload.html')

@app.route('/settings')
@admin_required
def settings(): return render_template('settings.html')

@app.route('/api/upload', methods=['POST'])
@admin_required
def api_upload():
    if 'file' not in request.files: return jsonify({'error':'No file'}), 400
    f = request.files['file']
    if not f.filename or f.filename.rsplit('.',1)[-1].lower() not in ALLOWED_EXTENSIONS:
        return jsonify({'error':'Invalid file type (.xlsx only)'}), 400
    filename = secure_filename(f.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    f.save(filepath)
    try:
        wb = load_workbook(filepath, read_only=True)
        required = ['Model_Inventory','Model_Validation','Validation_Checklist']
        missing = [s for s in required if s not in wb.sheetnames]
        if missing: return jsonify({'error':f'Missing sheets: {", ".join(missing)}. Found: {", ".join(wb.sheetnames)}'}), 400
        _rebuild_db(filepath)
        conn = get_db(); c = conn.cursor()
        c.execute("INSERT INTO upload_log (filename,uploaded_by,status) VALUES (?,?,?)",
                  (filename, session.get('username','admin'), 'success'))
        conn.commit(); conn.close()
        return jsonify({'success':True,'message':f'Successfully imported {filename}'})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

def _rebuild_db(filepath):
    wb = load_workbook(filepath, read_only=True)
    conn = get_db(); c = conn.cursor()
    for tbl in ['model_inventory','model_validation','validation_checklist']:
        c.execute(f"DROP TABLE IF EXISTS {tbl}")
    for sname, tname in [('Model_Inventory','model_inventory'),
                          ('Model_Validation','model_validation'),
                          ('Validation_Checklist','validation_checklist')]:
        if sname not in wb.sheetnames: continue
        ws = wb[sname]; rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        hdrs = [sanitize(h) for h in rows[0]]
        c.execute(f'CREATE TABLE {tname} ({", ".join(f"{h!r} TEXT" for h in hdrs)})')
        for row in rows[1:]:
            vals = [str(v) if v is not None else None for v in row]
            c.execute(f'INSERT INTO {tname} VALUES ({",".join(["?"]*len(hdrs))})', vals)
    # Ensure scoring columns exist in model_inventory for Parameter Scoring
    c.execute("PRAGMA table_info(model_inventory)")
    existing = [r[1] for r in c.fetchall()]
    scoring_cols = [
        "Materiality_Financial_Exposure", "Materiality_Customer_Exposure", "Materiality_Reputational", "Materiality_Sensitivity",
        "Criticality_Usage", "Criticality_User_Base", "Criticality_Usage_Frequency", "Criticality_Decision_Type",
        "Complexity_Methodology", "Complexity_Dependency", "Complexity_Input_Data", "Complexity_Product_Coverage", "Complexity_Implementation"
    ]
    for sc in scoring_cols:
        if sc not in existing:
            c.execute(f'ALTER TABLE model_inventory ADD COLUMN {sc} TEXT')
            
    conn.commit(); conn.close()

@app.route('/api/upload/history')
@admin_required
def api_upload_history():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM upload_log ORDER BY uploaded_at DESC LIMIT 20")
    rows = [dict(r) for r in c.fetchall()]; conn.close()
    return jsonify(rows)

@app.route('/api/has-data')
@login_required
def api_has_data(): return jsonify({'has_data': db_has_data()})

@app.route('/api/me')
@login_required
def api_me(): return jsonify({'username':session.get('username'),'role':session.get('role'),'full_name':session.get('full_name')})

# ── Settings ────────────────────────────────────────────────────────────────
@app.route('/api/settings/users')
@admin_required
def api_settings_users():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id,username,role,full_name,email,created_at,is_active FROM users ORDER BY id")
    rows = [dict(r) for r in c.fetchall()]; conn.close(); return jsonify(rows)

@app.route('/api/settings/users/create', methods=['POST'])
@admin_required
def api_create_user():
    d = request.json or {}
    u,p,r = d.get('username','').strip(), d.get('password','').strip(), d.get('role','user')
    if not u or not p: return jsonify({'error':'Username and password required'}), 400
    if r not in ('admin','developer','user'): return jsonify({'error':'Invalid role'}), 400
    try:
        conn = get_db(); c = conn.cursor()
        c.execute("INSERT INTO users (username,password_hash,role,full_name,email) VALUES (?,?,?,?,?)",
                  (u, hash_pw(p), r, d.get('full_name','').strip(), d.get('email','').strip()))
        conn.commit(); conn.close(); return jsonify({'success':True})
    except: return jsonify({'error':f'Username "{u}" already exists'}), 400

@app.route('/api/settings/users/<int:uid>/role', methods=['PUT'])
@admin_required
def api_change_role(uid):
    d = request.json or {}; new_role = d.get('role','')
    if new_role not in ('admin','developer','user'): return jsonify({'error':'Invalid role'}), 400
    if uid == session.get('user_id'): return jsonify({'error':'Cannot change own role'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("UPDATE users SET role=? WHERE id=?", (new_role, uid))
    conn.commit(); conn.close(); return jsonify({'success':True})

@app.route('/api/settings/users/<int:uid>/toggle', methods=['PUT'])
@admin_required
def api_toggle_user(uid):
    if uid == session.get('user_id'): return jsonify({'error':'Cannot deactivate yourself'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("UPDATE users SET is_active=CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE id=?", (uid,))
    conn.commit()
    c.execute("SELECT is_active FROM users WHERE id=?", (uid,))
    row = c.fetchone(); conn.close()
    return jsonify({'success':True,'is_active':row[0] if row else 0})

@app.route('/api/settings/users/<int:uid>/password', methods=['PUT'])
@admin_required
def api_reset_password(uid):
    d = request.json or {}; new_pw = d.get('password','').strip()
    if not new_pw or len(new_pw)<4: return jsonify({'error':'Min 4 characters'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("UPDATE users SET password_hash=? WHERE id=?", (hash_pw(new_pw), uid))
    conn.commit(); conn.close(); return jsonify({'success':True})

@app.route('/api/settings/users/<int:uid>', methods=['DELETE'])
@admin_required
def api_delete_user(uid):
    if uid == session.get('user_id'): return jsonify({'error':'Cannot delete yourself'}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM users WHERE id=?", (uid,)); conn.commit(); conn.close()
    return jsonify({'success':True})

# ── Overview ────────────────────────────────────────────────────────────────
@app.route('/api/overview/kpis')
@login_required
def api_overview_kpis():
    conn = get_db(); c = conn.cursor()
    # Get actual columns to build safe queries
    c.execute("PRAGMA table_info(model_inventory)"); cols = {r[1] for r in c.fetchall()}
    c.execute("SELECT COUNT(*) FROM model_inventory WHERE Model_ID IS NOT NULL AND Model_ID!='None' AND Model_ID!=''"); total = c.fetchone()[0]
    in_use = 0; ai = 0; validated = 0
    if 'Model_Status' in cols:
        c.execute("SELECT COUNT(*) FROM model_inventory WHERE Model_Status='Production'"); in_use = c.fetchone()[0]
    if 'AI_Model' in cols:
        c.execute("SELECT COUNT(*) FROM model_inventory WHERE AI_Model='Yes'"); ai = c.fetchone()[0]
    if 'Status_BE' in cols:
        c.execute("SELECT COUNT(*) FROM model_inventory WHERE Status_BE='Validated'"); validated = c.fetchone()[0]
    conn.close()
    return jsonify({'total':total,'in_use_pct':round(in_use/total*100,1) if total else 0,
                    'ai_pct':round(ai/total*100,1) if total else 0,
                    'production_pct':round(in_use/total*100,1) if total else 0,'validated':validated})

@app.route('/api/overview/tiering')
@login_required
def api_overview_tiering():
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_inventory)"); cols = {r[1] for r in c.fetchall()}
    rows = []
    if 'Model_Tiering' in cols:
        c.execute("SELECT Model_Tiering, COUNT(*) cnt FROM model_inventory WHERE Model_Tiering IS NOT NULL AND Model_Tiering!='None' GROUP BY Model_Tiering ORDER BY Model_Tiering")
        rows = [{'tier':r[0],'count':r[1]} for r in c.fetchall()]
    conn.close(); return jsonify(rows)

@app.route('/api/overview/family')
@login_required
def api_overview_family():
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_inventory)"); cols = {r[1] for r in c.fetchall()}
    # Support both old (Model_Family) and new (Risk_Type) column names
    family_col = 'Risk_Type' if 'Risk_Type' in cols else 'Model_Family'
    c.execute(f"SELECT {family_col}, COUNT(*) cnt FROM model_inventory WHERE {family_col} IS NOT NULL AND {family_col}!='None' GROUP BY {family_col} ORDER BY cnt DESC")
    family = [{'family':r[0],'count':r[1]} for r in c.fetchall()]
    c.execute("SELECT Model_Risk, COUNT(*) cnt FROM model_inventory WHERE Model_Risk IS NOT NULL AND Model_Risk!='None' GROUP BY Model_Risk ORDER BY cnt DESC")
    risk = [{'risk':r[0],'count':r[1]} for r in c.fetchall()]
    conn.close(); return jsonify({'family':family,'risk':risk})

@app.route('/api/overview/business-unit')
@login_required
def api_overview_bu():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT Business_Unit, COUNT(*) cnt FROM model_inventory WHERE Business_Unit IS NOT NULL AND Business_Unit!='None' GROUP BY Business_Unit ORDER BY cnt DESC")
    rows = [{'bu':r[0],'count':r[1]} for r in c.fetchall()]; conn.close(); return jsonify(rows)

@app.route('/api/overview/tiering-by-bu')
@login_required
def api_overview_tiering_by_bu():
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_inventory)"); cols = {r[1] for r in c.fetchall()}
    rows = []
    if 'Model_Tiering' in cols and 'Business_Unit' in cols:
        c.execute("""SELECT Business_Unit, Model_Tiering, COUNT(*) cnt
                     FROM model_inventory
                     WHERE Business_Unit IS NOT NULL AND Business_Unit!='None'
                     AND Model_Tiering IS NOT NULL AND Model_Tiering!='None'
                     GROUP BY Business_Unit, Model_Tiering
                     ORDER BY Business_Unit, Model_Tiering""")
        rows = [{'bu':r[0],'tier':r[1],'count':r[2]} for r in c.fetchall()]
    conn.close(); return jsonify(rows)

# ── Lifecycle ────────────────────────────────────────────────────────────────
@app.route('/api/lifecycle/models')
@login_required
def api_lifecycle_models():
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT * FROM model_inventory WHERE Model_ID IS NOT NULL AND Model_ID!='None' AND Model_ID!='' ORDER BY Model_Name""")
    rows = [dict(r) for r in c.fetchall()]; conn.close(); return jsonify(rows)

# ── Inventory ────────────────────────────────────────────────────────────────
@app.route('/api/inventory')
@login_required
def api_inventory():
    mid=request.args.get('model_id',''); fam=request.args.get('model_family',''); bu=request.args.get('business_unit','')
    conn = get_db(); c = conn.cursor()
    # Get actual columns present
    c.execute("PRAGMA table_info(model_inventory)"); actual_cols = {r[1] for r in c.fetchall()}
    def pick(*candidates):
        for col in candidates:
            if col in actual_cols: return col
        return None
    sox_col      = pick('SOX_Complient','SOX_Compliant','SOX_Compliant_')
    vendor_col   = pick('Vendor_Model','Vendor_Models')
    ai_col       = pick('AI_Model','is_AI_Model')
    doc_col      = pick('Model_Documentation','Documentation')
    exp_col      = pick('Exposure_of_Model_Usage_INR_Crores','Exposure_of_Model_Usage_INR_Cr','Exposure')
    mat_col      = pick('Materiality_Rating','Model_Materiality_Rating')
    crit_col     = pick('Criticality_Rating','Model_Criticality_Rating')
    complex_col  = pick('Complexity_Rating','Model_Complexity_Rating')
    risk_score_col = pick('Model_Risk_Score','Risk_Score')

    q = """SELECT * FROM model_inventory WHERE Model_ID IS NOT NULL AND Model_ID!='None' AND Model_ID!=''"""
    params = []
    # Support both old Model_Family and new Risk_Type column names
    c.execute("PRAGMA table_info(model_inventory)"); inv_cols = {r[1] for r in c.fetchall()}
    fam_col = 'Risk_Type' if 'Risk_Type' in inv_cols else 'Model_Family'
    if mid: q+=" AND Model_ID=?"; params.append(mid)
    if fam: q+=f" AND {fam_col}=?"; params.append(fam)
    if bu:  q+=" AND Business_Unit=?"; params.append(bu)
    c.execute(q,params)
    raw_rows=[dict(r) for r in c.fetchall()]
    # Normalize rows so frontend always gets standardized keys
    rows = []
    for row in raw_rows:
        row['_SOX_Compliant']   = row.get(sox_col,'')    if sox_col    else ''
        row['_Vendor_Model']    = row.get(vendor_col,'') if vendor_col else ''
        row['_AI_Model']        = row.get(ai_col,'')     if ai_col     else ''
        row['_Documentation']   = row.get(doc_col,'')    if doc_col    else ''
        row['_Exposure']        = row.get(exp_col,'')    if exp_col    else ''
        row['_Materiality']     = row.get(mat_col,'')    if mat_col    else ''
        row['_Criticality']     = row.get(crit_col,'')   if crit_col   else ''
        row['_Complexity']      = row.get(complex_col,'') if complex_col else ''
        row['_Risk_Score']      = row.get(risk_score_col,'') if risk_score_col else ''
        rows.append(row)
    conn.close(); return jsonify(rows)

@app.route('/api/inventory/filters')
@login_required
def api_inventory_filters():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT DISTINCT Model_ID, Model_Name FROM model_inventory WHERE Model_ID IS NOT NULL AND Model_ID!='' AND Model_ID!='None' ORDER BY Model_Name")
    models=[{'id':r[0], 'name':r[1]} for r in c.fetchall()]
    c.execute("PRAGMA table_info(model_inventory)"); fam_cols = {r[1] for r in c.fetchall()}
    family_col = 'Risk_Type' if 'Risk_Type' in fam_cols else 'Model_Family'
    c.execute(f"SELECT DISTINCT {family_col} FROM model_inventory WHERE {family_col} IS NOT NULL AND {family_col}!='None' ORDER BY {family_col}")
    fams=[r[0] for r in c.fetchall()]
    c.execute("SELECT DISTINCT Business_Unit FROM model_inventory WHERE Business_Unit IS NOT NULL AND Business_Unit!='None' ORDER BY Business_Unit")
    bus=[r[0] for r in c.fetchall()]
    conn.close(); return jsonify({'models':models,'families':fams,'business_units':bus,'family_col':family_col})

# ── Model Validation — FIXED ─────────────────────────────────────────────────
@app.route('/api/mv/kpis')
@login_required
def api_mv_kpis():
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_validation)"); mv_cols = {r[1] for r in c.fetchall()}
    # Use Model_Status if available, otherwise fall back to Status
    status_col = 'Model_Status' if 'Model_Status' in mv_cols else 'Status'
    # Total unique models
    c.execute("SELECT COUNT(DISTINCT Model_Name) FROM model_validation"); total = c.fetchone()[0]
    # Avg completion across all rows
    c.execute("SELECT AVG(CAST(REPLACE(Completion_pct,'%','') AS FLOAT)) FROM model_validation WHERE Completion_pct IS NOT NULL AND Completion_pct!='None'")
    avg_comp = c.fetchone()[0] or 0
    # Total findings (sum across all rows)
    c.execute("SELECT SUM(CAST(COALESCE(NULLIF(No_of_Findings,'None'),0) AS INTEGER)) FROM model_validation")
    findings = c.fetchone()[0] or 0
    # Not Started = models where ALL rows are Not Started (using Model_Status)
    c.execute(f"""SELECT COUNT(DISTINCT Model_Name) FROM model_validation
                 WHERE Model_Name NOT IN (
                   SELECT DISTINCT Model_Name FROM model_validation WHERE {status_col} != 'Not Started'
                 )""")
    not_started = c.fetchone()[0]
    # In Progress = models with at least one In Progress row (using Model_Status)
    c.execute(f"""SELECT COUNT(DISTINCT Model_Name) FROM model_validation WHERE {status_col}='In Progress'""")
    in_prog = c.fetchone()[0]
    # Completed count
    c.execute(f"""SELECT COUNT(DISTINCT Model_Name) FROM model_validation WHERE {status_col}='Completed'""")
    completed = c.fetchone()[0]
    conn.close()
    return jsonify({'total_models':total,'avg_completion':round(avg_comp,1),
                    'total_findings':int(findings),'not_started':not_started,'in_progress':in_prog,'completed':completed})

@app.route('/api/mv/matrix')
@login_required
def api_mv_matrix():
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT Risk_Category, Validation_Component, Sub_Step,
                 SUM(CASE WHEN Status='Completed' THEN 1 ELSE 0 END) as comp,
                 SUM(CASE WHEN Status='In Progress' THEN 1 ELSE 0 END) as ip,
                 SUM(CASE WHEN Status='Not Started' THEN 1 ELSE 0 END) as ns,
                 COUNT(*) as total,
                 ROUND(AVG(CAST(REPLACE(Completion_pct,'%','') AS FLOAT)),1) as avg_comp,
                 SUM(CAST(COALESCE(NULLIF(No_of_Findings,'None'),0) AS INTEGER)) as total_findings,
                 MIN(CASE WHEN Start_Date IS NOT NULL AND Start_Date!='None' THEN Start_Date END) as start_date,
                 MAX(CASE WHEN End_Date IS NOT NULL AND End_Date!='None' THEN End_Date END) as end_date,
                 GROUP_CONCAT(DISTINCT CASE WHEN Validator_Name IS NOT NULL AND Validator_Name!='None' THEN Validator_Name END) as validators
                 FROM model_validation
                 WHERE Risk_Category IS NOT NULL AND Risk_Category!='None'
                 GROUP BY Risk_Category, Validation_Component, Sub_Step
                 ORDER BY Risk_Category, Validation_Component, Sub_Step""")
    rows = []
    for r in c.fetchall():
        comp, ip, ns, total = r[3], r[4], r[5], r[6]
        # Dominant status
        if comp >= ip and comp >= ns: dom = 'Completed'
        elif ip >= ns: dom = 'In Progress'
        else: dom = 'Not Started'
        # Limit validator names to unique first 3
        validators_raw = r[11] or ''
        vlist = list(dict.fromkeys([v.strip() for v in validators_raw.split(',') if v.strip()]))[:3]
        vstr = ', '.join(vlist) + (f' +{len(validators_raw.split(","))-3} more' if len(validators_raw.split(',')) > 3 else '')
        rows.append({
            'Risk_Category':r[0],'Validation_Component':r[1],'Sub_Step':r[2],
            'dominant_status':dom,'comp':comp,'ip':ip,'ns':ns,'total':total,
            'avg_comp':r[7],'total_findings':r[8],'start_date':r[9],'end_date':r[10],
            'validators':vstr
        })
    conn.close(); return jsonify(rows)

@app.route('/api/mv/progress')
@login_required
def api_mv_progress():
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_validation)"); mv_cols = {r[1] for r in c.fetchall()}
    status_col = 'Model_Status' if 'Model_Status' in mv_cols else 'Status'
    c.execute("""SELECT Validation_Component,
                 ROUND(AVG(CAST(REPLACE(Completion_pct,'%','') AS FLOAT)),1) pct
                 FROM model_validation WHERE Validation_Component IS NOT NULL
                 GROUP BY Validation_Component ORDER BY pct DESC""")
    by_comp=[{'component':r[0],'pct':r[1]} for r in c.fetchall()]
    c.execute("""SELECT Risk_Category,
                 ROUND(AVG(CAST(REPLACE(Completion_pct,'%','') AS FLOAT)),1) pct
                 FROM model_validation WHERE Risk_Category IS NOT NULL AND Risk_Category!='None'
                 GROUP BY Risk_Category ORDER BY pct DESC""")
    by_risk=[{'risk':r[0],'pct':r[1]} for r in c.fetchall()]
    c.execute(f"SELECT {status_col}, COUNT(*) cnt FROM model_validation WHERE {status_col} IS NOT NULL AND {status_col}!='None' GROUP BY {status_col} ORDER BY cnt DESC")
    status_dist=[{'status':r[0],'count':r[1]} for r in c.fetchall()]
    conn.close(); return jsonify({'by_component':by_comp,'by_risk':by_risk,'status_dist':status_dist})

@app.route('/api/mv/findings')
@login_required
def api_mv_findings():
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_validation)"); mv_cols = {r[1] for r in c.fetchall()}
    status_col = 'Model_Status' if 'Model_Status' in mv_cols else 'Status'
    # Stacked bar: Y=Sum of Findings, Grouped by Risk Category, Stacked by Status
    c.execute(f"""SELECT Risk_Category,
                 SUM(CAST(COALESCE(NULLIF(No_of_Findings,'None'),0) AS INTEGER)) as total_findings,
                 SUM(CASE WHEN {status_col}='Completed' THEN CAST(COALESCE(NULLIF(No_of_Findings,'None'),0) AS INTEGER) ELSE 0 END) as comp_findings,
                 SUM(CASE WHEN {status_col}='In Progress' THEN CAST(COALESCE(NULLIF(No_of_Findings,'None'),0) AS INTEGER) ELSE 0 END) as ip_findings,
                 SUM(CASE WHEN {status_col}='Not Started' THEN CAST(COALESCE(NULLIF(No_of_Findings,'None'),0) AS INTEGER) ELSE 0 END) as ns_findings
                 FROM model_validation WHERE Risk_Category IS NOT NULL AND Risk_Category!='None'
                 GROUP BY Risk_Category ORDER BY total_findings DESC""")
    rows=[dict(r) for r in c.fetchall()]; conn.close(); return jsonify(rows)

@app.route('/api/mv/timeline')
@login_required
def api_mv_timeline():
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_validation)"); mv_cols = {r[1] for r in c.fetchall()}
    # Use Model_Status column directly if available, else fall back to deriving from Completion_pct
    if 'Model_Status' in mv_cols:
        c.execute("""SELECT Model_Name,
                     MIN(CASE WHEN Start_Date IS NOT NULL AND Start_Date NOT IN ('None','','nan') THEN Start_Date END) as Start_Date,
                     MAX(CASE WHEN End_Date   IS NOT NULL AND End_Date   NOT IN ('None','','nan') THEN End_Date   END) as End_Date,
                     ROUND(AVG(CAST(REPLACE(COALESCE(NULLIF(Completion_pct,'None'),'0'),'%','') AS FLOAT)),1) as avg_pct,
                     MAX(CASE
                       WHEN Model_Status='Completed' THEN 'Completed'
                       WHEN Model_Status='In Progress' THEN 'In Progress'
                       ELSE 'Not Started'
                     END) as Status
                     FROM model_validation
                     GROUP BY Model_Name
                     ORDER BY Start_Date""")
    else:
        c.execute("""SELECT Model_Name,
                     MIN(CASE WHEN Start_Date IS NOT NULL AND Start_Date NOT IN ('None','','nan') THEN Start_Date END) as Start_Date,
                     MAX(CASE WHEN End_Date   IS NOT NULL AND End_Date   NOT IN ('None','','nan') THEN End_Date   END) as End_Date,
                     ROUND(AVG(CAST(REPLACE(COALESCE(NULLIF(Completion_pct,'None'),'0'),'%','') AS FLOAT)),1) as avg_pct,
                     CASE
                       WHEN AVG(CAST(REPLACE(COALESCE(NULLIF(Completion_pct,'None'),'0'),'%','') AS FLOAT)) >= 95 THEN 'Completed'
                       WHEN AVG(CAST(REPLACE(COALESCE(NULLIF(Completion_pct,'None'),'0'),'%','') AS FLOAT)) > 0   THEN 'In Progress'
                       ELSE 'Not Started'
                     END as Status
                     FROM model_validation
                     GROUP BY Model_Name
                     ORDER BY Start_Date""")
    rows=[dict(r) for r in c.fetchall()]; conn.close(); return jsonify(rows)

@app.route('/api/mv/models')
@login_required
def api_mv_models():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT DISTINCT Model_Name FROM model_validation WHERE Model_Name IS NOT NULL AND Model_Name!='None' ORDER BY Model_Name")
    models=[r[0] for r in c.fetchall()]; conn.close(); return jsonify(models)

@app.route('/api/mv/matrix_by_model')
@login_required
def api_mv_matrix_by_model():
    model = request.args.get('model','')
    conn = get_db(); c = conn.cursor()
    where = "WHERE Risk_Category IS NOT NULL AND Risk_Category!='None'"
    params = []
    if model:
        where += " AND TRIM(Model_Name)=?"
        params.append(model.strip())
    c.execute(f"""SELECT Risk_Category, Validation_Component, Sub_Step,
                 SUM(CASE WHEN Status='Completed' THEN 1 ELSE 0 END) as comp,
                 SUM(CASE WHEN Status='In Progress' THEN 1 ELSE 0 END) as ip,
                 SUM(CASE WHEN Status='Not Started' THEN 1 ELSE 0 END) as ns,
                 COUNT(*) as total,
                 ROUND(AVG(CAST(REPLACE(Completion_pct,'%','') AS FLOAT)),1) as avg_comp,
                 SUM(CAST(COALESCE(NULLIF(No_of_Findings,'None'),0) AS INTEGER)) as total_findings,
                 MIN(CASE WHEN Start_Date IS NOT NULL AND Start_Date NOT IN ('None','') THEN Start_Date END) as start_date,
                 MAX(CASE WHEN End_Date   IS NOT NULL AND End_Date   NOT IN ('None','') THEN End_Date   END) as end_date,
                 GROUP_CONCAT(DISTINCT CASE WHEN Validator_Name IS NOT NULL AND Validator_Name!='None' THEN Validator_Name END) as validators
                 FROM model_validation {where}
                 GROUP BY Risk_Category, Validation_Component, Sub_Step
                 ORDER BY Risk_Category, Validation_Component, Sub_Step""", params)
    rows = []
    for r in c.fetchall():
        comp, ip, ns, total = r[3], r[4], r[5], r[6]
        if comp >= ip and comp >= ns: dom = 'Completed'
        elif ip >= ns: dom = 'In Progress'
        else: dom = 'Not Started'
        validators_raw = r[11] or ''
        vlist = list(dict.fromkeys([v.strip() for v in validators_raw.split(',') if v.strip()]))[:3]
        extra = len([v for v in validators_raw.split(',') if v.strip()]) - 3
        vstr = ', '.join(vlist) + (f' +{extra} more' if extra > 0 else '')
        rows.append({
            'Risk_Category':r[0],'Validation_Component':r[1],'Sub_Step':r[2],
            'dominant_status':dom,'comp':comp,'ip':ip,'ns':ns,'total':total,
            'avg_comp':r[7],'total_findings':r[8],'start_date':r[9],'end_date':r[10],
            'validators':vstr
        })
    conn.close(); return jsonify(rows)

# ── Executive ────────────────────────────────────────────────────────────────
@app.route('/api/executive/kpis')
@login_required
def api_executive_kpis():
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_inventory)"); cols = {r[1] for r in c.fetchall()}
    c.execute("SELECT COUNT(*) FROM model_inventory WHERE Model_ID IS NOT NULL AND Model_ID!='None' AND Model_ID!=''"); total=c.fetchone()[0]
    validated=in_prog=overdue=ns=0; comp=0
    if 'Status_BE' in cols:
        c.execute("SELECT COUNT(*) FROM model_inventory WHERE Status_BE='Validated'"); validated=c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM model_inventory WHERE Status_BE='In Progress'"); in_prog=c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM model_inventory WHERE Status_BE='Overdue'"); overdue=c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM model_inventory WHERE Status_BE='Not Started'"); ns=c.fetchone()[0]
    if 'Overall_pct' in cols:
        c.execute("SELECT AVG(CAST(REPLACE(Overall_pct,'%','') AS FLOAT)) FROM model_inventory WHERE Overall_pct IS NOT NULL AND Overall_pct!='None' AND Overall_pct!=''")
        raw=c.fetchone()[0] or 0
        # Values stored as decimals (0-1) need to be converted to percentage (0-100)
        comp = raw * 100 if raw <= 1.5 else raw
    conn.close()
    return jsonify({'total':total,'validated':validated,'in_progress':in_prog,'overdue':overdue,'not_started':ns,'completion':round(comp,1)})

@app.route('/api/executive/table')
@login_required
def api_executive_table():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM model_inventory WHERE Model_ID IS NOT NULL AND Model_ID!='None' AND Model_ID!='' ORDER BY Model_Name")
    rows=[dict(r) for r in c.fetchall()]; conn.close(); return jsonify(rows)

# ── Timeline (from Model_Inventory - ALL models with date fallback) ──────────
@app.route('/api/timeline')
@login_required
def api_timeline():
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_inventory)"); cols = {r[1] for r in c.fetchall()}
    # Build date coalesce dynamically based on available columns
    date_cols_start = [c2 for c2 in ['Start_Date','Identification_Date','Last_Annual_Review_Date','Last_Review_Date'] if c2 in cols]
    date_cols_end   = [c2 for c2 in ['End_Date','Validation_Date','Due_Date','Next_Annual_Review_Date','Next_Review_Date'] if c2 in cols]
    status_col = 'Status_BE' if 'Status_BE' in cols else ('Model_Status' if 'Model_Status' in cols else None)
    identified_col = 'Identified_By' if 'Identified_By' in cols else None

    def coalesce_dates(col_list, tbl='mi'):
        parts = [f"NULLIF({tbl}.{c2},'None')" for c2 in col_list]
        # fallback to model_validation
        parts.append(f"(SELECT MIN(mv.Start_Date) FROM model_validation mv WHERE mv.Model_Name={tbl}.Model_Name AND mv.Start_Date IS NOT NULL AND mv.Start_Date!='None')" if col_list == date_cols_start else f"(SELECT MAX(mv.End_Date) FROM model_validation mv WHERE mv.Model_Name={tbl}.Model_Name AND mv.End_Date IS NOT NULL AND mv.End_Date!='None')")
        return f"COALESCE({', '.join(parts)})"

    sd_expr = coalesce_dates(date_cols_start) if date_cols_start else "(SELECT MIN(mv.Start_Date) FROM model_validation mv WHERE mv.Model_Name=mi.Model_Name AND mv.Start_Date IS NOT NULL AND mv.Start_Date!='None')"
    ed_expr = coalesce_dates(date_cols_end)   if date_cols_end   else "(SELECT MAX(mv.End_Date) FROM model_validation mv WHERE mv.Model_Name=mi.Model_Name AND mv.End_Date IS NOT NULL AND mv.End_Date!='None')"
    status_expr = f'mi.{status_col}' if status_col else "'Unknown'"
    identified_expr = f'mi.{identified_col}' if identified_col else "NULL"

    # Detect Risk_Type vs Model_Family
    family_col_tl = 'Risk_Type' if 'Risk_Type' in cols else 'Model_Family'
    c.execute(f"""SELECT mi.Model_Name, mi.{family_col_tl} as Risk_Type, mi.{family_col_tl} as Model_Family,
                 {identified_expr} as Identified_By, {status_expr} as Status_BE,
                 {sd_expr} as Start_Date,
                 {ed_expr} as End_Date
                 FROM model_inventory mi
                 ORDER BY Start_Date""")
    rows=[dict(r) for r in c.fetchall()]; conn.close(); return jsonify(rows)

# ── Validation Checklist ─────────────────────────────────────────────────────
@app.route('/api/checklist')
@login_required
def api_checklist():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT DISTINCT Model_Name FROM validation_checklist ORDER BY Model_Name")
    models=[r[0] for r in c.fetchall()]
    # Use DISTINCT to prevent inflated counts if the same data was uploaded multiple times
    c.execute("""SELECT DISTINCT Validation_Part, Validation_Section, Substep_Name, Model_Name, Status 
                 FROM validation_checklist 
                 ORDER BY Validation_Part, Validation_Section, Substep_Code""")
    rows=[dict(r) for r in c.fetchall()]; conn.close()
    return jsonify({'models':models,'data':rows})

@app.route('/parameter-scoring')
@privileged_required
def parameter_scoring(): return render_template('parameter_scoring.html')

# ── Model Detail Page ─────────────────────────────────────────────────────────
@app.route('/models/<model_id>')
@login_required
def model_detail(model_id):
    return render_template('model_detail.html', model_id=model_id)

@app.route('/api/inventory/<model_id>')
@login_required
def api_inventory_detail(model_id):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM model_inventory WHERE Model_ID=?", (model_id,))
    row = c.fetchone()
    if not row: conn.close(); return jsonify({'error':'Not found'}), 404
    data = dict(row)
    # Get validation records from model_validation
    # Aggressive Grouping: One card per Sub_Step to prevent any duplication (94/68 vs 47)
    m_name = data.get('Model_Name','')
    c.execute("""SELECT * FROM (
                    SELECT rowid, * FROM model_validation 
                    WHERE TRIM(Model_ID)=? OR Model_Name=? OR Model_Name LIKE ?
                    ORDER BY Completion_pct DESC
                 ) GROUP BY Validation_Component, Sub_Step""",
              (model_id.strip(), m_name, f"%{m_name[:15]}%"))
    val_rows = [dict(r) for r in c.fetchall()]

    # Get checklist
    c.execute("""SELECT * FROM (
                    SELECT rowid, * FROM validation_checklist 
                    WHERE TRIM(Model_ID)=? OR Model_Name=? OR Model_Name LIKE ?
                 ) GROUP BY Validation_Part, Validation_Section, Substep_Name""",
              (model_id.strip(), m_name, f"%{m_name[:15]}%"))
    checklist_rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify({'model': data, 'validation_records': val_rows, 'checklist': checklist_rows})

@app.route('/api/inventory/<model_id>', methods=['PUT'])
@admin_required
def api_inventory_update(model_id):
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_inventory)")
    cols = [r[1] for r in c.fetchall()]
    updates = {k: v for k, v in d.items() if k in cols and k != 'Model_ID'}
    
    # Calculate scores if sub-parameters are provided
    sub_flds = [
        "Materiality_Financial_Exposure", "Materiality_Customer_Exposure", "Materiality_Reputational", "Materiality_Sensitivity",
        "Criticality_Usage", "Criticality_User_Base", "Criticality_Usage_Frequency", "Criticality_Decision_Type",
        "Complexity_Methodology", "Complexity_Dependency", "Complexity_Input_Data", "Complexity_Product_Coverage", "Complexity_Implementation"
    ]
    
    # Check if we need to recalculate
    if any(f in updates for f in sub_flds) or any(f in updates for f in ["Materiality_Rating", "Criticality_Rating", "Complexity_Rating"]):
        # Fetch current values to fill gaps
        c.execute("SELECT * FROM model_inventory WHERE Model_ID=?", (model_id,))
        curr = dict(c.fetchone())
        
        def gv(f):
            v = updates.get(f, curr.get(f))
            try: return float(v)
            except: return 1.0
            
        # Materiality: Fin(0.35), Cust(0.35), Rep(0.2), Sens(0.1)
        m_score = (gv('Materiality_Financial_Exposure') * 0.35 + 
                   gv('Materiality_Customer_Exposure') * 0.35 + 
                   gv('Materiality_Reputational') * 0.2 + 
                   gv('Materiality_Sensitivity') * 0.1)
        
        # Criticality: Usage(0.35), User(0.35), Freq(0.2), Decision(0.1)
        crit_score = (gv('Criticality_Usage') * 0.35 + 
                      gv('Criticality_User_Base') * 0.35 + 
                      gv('Criticality_Usage_Frequency') * 0.2 + 
                      gv('Criticality_Decision_Type') * 0.1)
        
        # Complexity: 0.2 each
        comp_score = (gv('Complexity_Methodology') + 
                      gv('Complexity_Dependency') + 
                      gv('Complexity_Input_Data') + 
                      gv('Complexity_Product_Coverage') + 
                      gv('Complexity_Implementation')) * 0.2
                      
        updates['Materiality_Rating'] = round(m_score, 2)
        updates['Criticality_Rating'] = round(crit_score, 2)
        updates['Complexity_Rating'] = round(comp_score, 2)
        
        # Overall Risk Score: Mat(0.4), Crit(0.4), Comp(0.2)
        risk_score = (m_score * 0.4 + crit_score * 0.4 + comp_score * 0.2)
        updates['Model_Risk_Score'] = round(risk_score, 2)
        
        # Tiering: T1[2.2, 3], T2[1.8, 2.2), T3[1, 1.8)
        if risk_score >= 2.2: tier = 'Tier 1'
        elif risk_score >= 1.8: tier = 'Tier 2'
        else: tier = 'Tier 3'
        updates['Model_Tiering'] = tier
        updates['Model_Risk_Tier'] = tier

    if not updates: conn.close(); return jsonify({'error':'No valid fields'}), 400
    
    set_clause = ', '.join(f'"{k}"=?' for k in updates)
    vals = list(updates.values()) + [model_id]
    c.execute(f"UPDATE model_inventory SET {set_clause} WHERE Model_ID=?", vals)
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/inventory/<model_id>', methods=['DELETE'])
@admin_required
def api_inventory_delete(model_id):
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM model_inventory WHERE Model_ID=?", (model_id,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/inventory', methods=['POST'])
@admin_required
def api_inventory_create():
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_inventory)")
    cols = [r[1] for r in c.fetchall()]
    inserts = {k: v for k, v in d.items() if k in cols}
    if not inserts or 'Model_ID' not in inserts:
        conn.close(); return jsonify({'error':'Model_ID required'}), 400
    keys = list(inserts.keys()); vals = list(inserts.values())
    col_str = ', '.join('"' + k + '"' for k in keys)
    ph_str = ', '.join(['?'] * len(vals))
    c.execute("INSERT INTO model_inventory (" + col_str + ") VALUES (" + ph_str + ")", vals)
    conn.commit(); conn.close()
    return jsonify({'success': True})

# ── Model Validation CRUD ─────────────────────────────────────────────────────
@app.route('/api/mv/row/<int:rowid>', methods=['PUT'])
@admin_required
def api_mv_update(rowid):
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_validation)")
    cols = [r[1] for r in c.fetchall()]
    updates = {k: v for k, v in d.items() if k in cols}
    if not updates: conn.close(); return jsonify({'error':'No valid fields'}), 400
    set_clause = ', '.join(f'"{k}"=?' for k in updates)
    vals = list(updates.values()) + [rowid]
    c.execute(f"UPDATE model_validation SET {set_clause} WHERE rowid=?", vals)
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/mv/row/<int:rowid>', methods=['DELETE'])
@admin_required
def api_mv_delete(rowid):
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM model_validation WHERE rowid=?", (rowid,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/mv/row', methods=['POST'])
@admin_required
def api_mv_create():
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(model_validation)")
    cols = [r[1] for r in c.fetchall()]
    inserts = {k: v for k, v in d.items() if k in cols}
    if not inserts: conn.close(); return jsonify({'error':'No data'}), 400
    keys = list(inserts.keys()); vals = list(inserts.values())
    col_str = ', '.join('"' + k + '"' for k in keys)
    ph_str = ', '.join(['?'] * len(vals))
    c.execute("INSERT INTO model_validation (" + col_str + ") VALUES (" + ph_str + ")", vals)
    conn.commit(); conn.close()
    return jsonify({'success': True})

# ── Validation Checklist CRUD ──────────────────────────────────────────────────
@app.route('/api/checklist/row/<int:rowid>', methods=['PUT'])
@admin_required
def api_checklist_update(rowid):
    d = request.json or {}
    conn = get_db(); c = conn.cursor()
    c.execute("PRAGMA table_info(validation_checklist)")
    cols = [r[1] for r in c.fetchall()]
    updates = {k: v for k, v in d.items() if k in cols}
    if not updates: conn.close(); return jsonify({'error':'No valid fields'}), 400
    set_clause = ', '.join(f'"{k}"=?' for k in updates)
    vals = list(updates.values()) + [rowid]
    c.execute(f"UPDATE validation_checklist SET {set_clause} WHERE rowid=?", vals)
    conn.commit(); conn.close()
    return jsonify({'success': True})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
